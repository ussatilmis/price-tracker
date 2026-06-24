"""
Paylasilabilir Excel rapor uretir:
  - 'Guncel Fiyatlar' sayfasi: tum urunler, para birimi formatli
  - 'Fiyat Degisimleri' sayfasi: dususler yesil, artislar kirmizi
Musteriye teslim edilen sey genelde bu .xlsx olur; CSV'ler motorun arkasinda durur.
"""

import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
DROP_FILL = PatternFill("solid", fgColor="C6EFCE")   # fiyat dustu -> yesil
RISE_FILL = PatternFill("solid", fgColor="FFC7CE")   # fiyat artti -> kirmizi
NEW_FILL = PatternFill("solid", fgColor="FFF2CC")    # yeni urun -> sari
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _read_csv(path):
    if not os.path.exists(path):
        return []
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _style_header(ws, headers):
    for col, name in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill, c.font = HEADER_FILL, HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.freeze_panes = "A2"


def _autosize(ws, widths):
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _sheet_latest(ws, rows):
    headers = ["Product", "Price", "In Stock", "URL"]
    _style_header(ws, headers)
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r.get("title")).border = BORDER
        pc = ws.cell(row=i, column=2, value=_to_float(r.get("price")))
        pc.number_format = "#,##0.00"
        pc.border = BORDER
        ws.cell(row=i, column=3, value="Yes" if r.get("in_stock") == "True" else "No").border = BORDER
        ws.cell(row=i, column=4, value=r.get("url")).border = BORDER
    _autosize(ws, [55, 12, 10, 60])


def _sheet_changes(ws, rows):
    headers = ["Date", "Product", "Status", "Old Price", "New Price", "Change"]
    _style_header(ws, headers)
    # en yeni olaylar ustte
    rows = sorted(rows, key=lambda r: r.get("scraped_at", ""), reverse=True)
    for i, r in enumerate(rows, start=2):
        change = r.get("change")
        delta = _to_float(r.get("delta"))
        fill = NEW_FILL if change == "NEW" else (DROP_FILL if (delta or 0) < 0 else RISE_FILL)
        values = [
            r.get("scraped_at"),
            r.get("title"),
            {"NEW": "NEW", "PRICE_CHANGE": "PRICE CHANGE"}.get(change, change),
            _to_float(r.get("old_price")),
            _to_float(r.get("price")),
            delta,
        ]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.border = BORDER
            c.fill = fill
            if col in (4, 5, 6):
                c.number_format = "#,##0.00;[Red]-#,##0.00"
    _autosize(ws, [22, 55, 16, 12, 12, 10])


def build_report():
    wb = Workbook()
    _sheet_latest(wb.active, _read_csv(config.LATEST_CSV))
    wb.active.title = "Guncel Fiyatlar"
    _sheet_changes(wb.create_sheet("Fiyat Degisimleri"), _read_csv(config.HISTORY_CSV))
    os.makedirs(config.DATA_DIR, exist_ok=True)
    wb.save(config.REPORT_XLSX)
    return config.REPORT_XLSX
